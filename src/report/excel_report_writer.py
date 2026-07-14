from datetime import datetime
from pathlib import Path

from openpyxl import Workbook


class ExcelReportWriter:

    def write(
        self,
        count_result,
        sum_result,
        groupby_result,
        rowcompare_result
    ):

        output_dir = Path("output")
        output_dir.mkdir(exist_ok=True)

        file_name = (
            f"validation_result_"
            f"{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        )

        file_path = output_dir / file_name

        workbook = Workbook()

        summary_sheet = workbook.active
        summary_sheet.title = "Summary"

        sum_sheet = workbook.create_sheet("Sum Validation")
        groupby_sheet = workbook.create_sheet("Group By Validation")
        rowcompare_sheet = workbook.create_sheet("Row Compare Validation")

        # Summary
        summary_sheet.append(["Validation", "Result"])

        summary_sheet.append([
            "Count Validation",
            "PASS" if count_result else "FAIL"
        ])

        summary_sheet.append([
            "Sum Validation",
            "PASS" if sum(
                1
                for result in sum_result
                if not result["result"]
            ) == 0 else "FAIL"
        ])

        summary_sheet.append([
            "Group By Validation",
            "PASS" if groupby_result["result"] else "FAIL"
        ])

        summary_sheet.append([
            "Row Compare Validation",
            "PASS" if rowcompare_result["result"] else "FAIL"
        ])

        # Sum Validation
        sum_sheet.append([
            "Column",
            "Source",
            "Target",
            "Result"
        ])

        for result in sum_result:

            sum_sheet.append([
                result["column"],
                result["source_sum"],
                result["target_sum"],
                "PASS" if result["result"] else "FAIL"
            ])

        # Group By Validation
        headers = (
            groupby_result["group_by_columns"]
            + [
                f"{column}_source"
                for column in groupby_result["compare_columns"]
            ]
            + [
                f"{column}_target"
                for column in groupby_result["compare_columns"]
            ]
            + [
                f"{column}_result"
                for column in groupby_result["compare_columns"]
            ]
        )

        groupby_sheet.append(headers)

        for _, row in groupby_result["difference_df"].iterrows():

            values = []

            for column in groupby_result["group_by_columns"]:
                values.append(row[column])

            for column in groupby_result["compare_columns"]:
                values.append(row[f"{column}_source"])

            for column in groupby_result["compare_columns"]:
                values.append(row[f"{column}_target"])

            for column in groupby_result["compare_columns"]:
                values.append(
                    "PASS"
                    if row[f"{column}_result"]
                    else "FAIL"
                )

            groupby_sheet.append(values)

        # Row Compare Validation
        rowcompare_sheet.append([
            "Type",
            "PK",
            "Column",
            "Source",
            "Target"
        ])

        for update in rowcompare_result["update_rows"]:

            for change in update["changes"]:

                rowcompare_sheet.append([
                    "UPDATE",
                    update["pk"],
                    change["column"],
                    change["source"],
                    change["target"]
                ])

        for _, row in rowcompare_result["delete_rows"].iterrows():

            rowcompare_sheet.append([
                "DELETE",
                row["sale_id"],
                "",
                "",
                ""
            ])

        for _, row in rowcompare_result["insert_rows"].iterrows():

            rowcompare_sheet.append([
                "INSERT",
                row["sale_id"],
                "",
                "",
                ""
            ])

        workbook.save(file_path)

        return file_path